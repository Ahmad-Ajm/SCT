"""
exporters/excel_exporter.py
============================
تصدير كل البيانات في ملف Excel واحد متعدد الأوراق.

هذا هو الملف الأهم للتسليم للعميل.

يحتوي على:
- Overview (ملخص)
- Critical/High/Medium/Low Issues
- Pages, Links, Images, Schema, etc.
- مع تنسيق احترافي وألوان حسب الأولوية
"""

from dataclasses import asdict, is_dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils.logger import get_logger

log = get_logger(__name__)


# === ألوان موحّدة ===
COLORS = {
    "header_bg": "1F4E79",  # أزرق داكن
    "header_text": "FFFFFF",
    "critical": "FFC7CE",  # أحمر فاتح
    "high": "FFEB9C",  # برتقالي فاتح
    "medium": "FFF2CC",  # أصفر فاتح
    "low": "E2EFDA",  # أخضر فاتح
    "alt_row": "F2F2F2",  # رمادي فاتح للصفوف البديلة
}


class ExcelExporter:
    """
    مصدّر Excel موحّد.

    Example:
        >>> exporter = ExcelExporter("./output", filename="audit.xlsx")
        >>> exporter.export(crawl_data, analysis_data)
    """

    def __init__(self, output_dir: str, filename: str = "master_audit.xlsx"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.file_path = self.output_dir / filename

    def export(
        self,
        pages: list[Any],
        links: list[dict[str, Any]],
        images: list[dict[str, Any]],
        headings: list[dict[str, Any]],
        schema: list[dict[str, Any]],
        redirects: list[dict[str, Any]],
        headers: list[dict[str, Any]],
        seo_issues: dict[str, Any],
        duplicate_data: dict[str, Any],
        orphan_data: dict[str, Any],
        thin_content_data: dict[str, Any],
        broken_data: dict[str, Any],
        images_analysis: dict[str, Any],
        crawl_stats: Any = None,
        site_url: str = "",
    ) -> str:
        """
        إنشاء ملف Excel موحّد.

        Returns:
            str: مسار الملف
        """
        log.info(f"بدء إنشاء Excel: {self.file_path}")

        wb = Workbook()
        # إزالة الورقة الافتراضية
        wb.remove(wb.active)

        # === Sheet 1: Overview ===
        self._create_overview_sheet(
            wb, pages, links, images, schema, seo_issues, crawl_stats, site_url
        )

        # === Sheet 2-5: Issues by Severity ===
        self._create_issues_sheets(wb, seo_issues)

        # === Sheet 6: Pages ===
        pages_dicts = [self._to_dict(p) for p in pages]
        self._create_data_sheet(wb, "Pages", pages_dicts, important_columns=[
            "url", "status_code", "title", "title_length", "meta_description",
            "h1_count", "word_count", "is_indexable", "indexability_reason", "depth"
        ])

        # === Sheet 7: Links ===
        self._create_data_sheet(wb, "All Links", links, important_columns=[
            "from_url", "to_url", "anchor_text", "is_internal", "nofollow", "in_navigation"
        ])

        # === Sheet 8: Images ===
        self._create_data_sheet(wb, "Images", images, important_columns=[
            "page_url", "src", "alt", "has_alt", "has_explicit_dimensions", "file_extension"
        ])

        # === Sheet 9: Schema ===
        schema_flat = [
            {
                "page_url": item.get("page_url", ""),
                "format": item.get("format", ""),
                "type": item.get("type", ""),
                "name": item.get("name", ""),
            }
            for item in schema
        ]
        self._create_data_sheet(wb, "Schema", schema_flat)

        # === Sheet 10: Redirects ===
        self._create_data_sheet(wb, "Redirects", redirects)

        # === Sheet 11: 404 Pages ===
        self._create_data_sheet(wb, "404 Pages", broken_data.get("pages_4xx", []))

        # === Sheet 12: Duplicate Titles ===
        dup_titles_flat = []
        for d in duplicate_data.get("duplicate_titles", []):
            for url in d["urls"]:
                dup_titles_flat.append({
                    "title": d["value"],
                    "url": url,
                    "duplicate_count": d["count"],
                })
        self._create_data_sheet(wb, "Duplicate Titles", dup_titles_flat)

        # === Sheet 13: Orphan Pages ===
        self._create_data_sheet(wb, "Orphan Pages", orphan_data.get("orphan_pages", []))

        # === Sheet 14: Thin Content ===
        thin_all = (
            thin_content_data.get("critical_thin_pages", []) +
            thin_content_data.get("thin_content_pages", [])
        )
        self._create_data_sheet(wb, "Thin Content", thin_all)

        # === حفظ ===
        try:
            wb.save(self.file_path)
            log.info(f"تم حفظ Excel بنجاح: {self.file_path}")
            return str(self.file_path)
        except Exception as e:
            log.error(f"فشل حفظ Excel: {e}")
            return ""

    # ========================================================
    # === Sheet Builders ===
    # ========================================================

    def _create_overview_sheet(
        self, wb, pages, links, images, schema, seo_issues, crawl_stats, site_url
    ):
        """إنشاء ورقة Overview."""
        ws = wb.create_sheet("📊 Overview", 0)

        # === Header ===
        ws["A1"] = f"تقرير تدقيق SEO — {site_url}"
        ws["A1"].font = Font(size=18, bold=True, color=COLORS["header_text"])
        ws["A1"].fill = PatternFill("solid", fgColor=COLORS["header_bg"])
        ws.merge_cells("A1:D1")
        ws.row_dimensions[1].height = 35

        from datetime import datetime
        ws["A2"] = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = Font(italic=True, size=10)

        # === الإحصائيات الأساسية ===
        row = 4
        ws.cell(row=row, column=1, value="📈 إحصائيات الزحف").font = Font(size=14, bold=True)
        row += 1

        stats_data = [
            ("إجمالي الصفحات المُزحوفة", len(pages)),
            ("الصفحات الناجحة (2xx)", sum(1 for p in pages if 200 <= self._get_attr(p, "status_code", 0) < 300)),
            ("Redirects (3xx)", sum(1 for p in pages if 300 <= self._get_attr(p, "status_code", 0) < 400)),
            ("Client Errors (4xx)", sum(1 for p in pages if 400 <= self._get_attr(p, "status_code", 0) < 500)),
            ("Server Errors (5xx)", sum(1 for p in pages if 500 <= self._get_attr(p, "status_code", 0) < 600)),
            ("صفحات قابلة للفهرسة", sum(1 for p in pages if self._get_attr(p, "is_indexable", False))),
            ("صفحات غير قابلة للفهرسة", sum(1 for p in pages if not self._get_attr(p, "is_indexable", True))),
            ("إجمالي الروابط الداخلية", sum(1 for l in links if l.get("is_internal"))),
            ("إجمالي الروابط الخارجية", sum(1 for l in links if not l.get("is_internal"))),
            ("إجمالي الصور", len(images)),
            ("إجمالي Schema entries", len(schema)),
        ]

        for label, value in stats_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        # === ملخص المشاكل ===
        row += 2
        ws.cell(row=row, column=1, value="🎯 ملخص المشاكل").font = Font(size=14, bold=True)
        row += 1

        summary = seo_issues.get("summary", {})
        issues_summary = [
            ("🔴 Critical", summary.get("critical_count", 0), COLORS["critical"]),
            ("🟠 High", summary.get("high_count", 0), COLORS["high"]),
            ("🟡 Medium", summary.get("medium_count", 0), COLORS["medium"]),
            ("🟢 Low", summary.get("low_count", 0), COLORS["low"]),
            ("الإجمالي", summary.get("total_issues", 0), None),
        ]

        for label, count, color in issues_summary:
            cell_label = ws.cell(row=row, column=1, value=label)
            cell_value = ws.cell(row=row, column=2, value=count)
            cell_label.font = Font(bold=True)
            if color:
                cell_label.fill = PatternFill("solid", fgColor=color)
                cell_value.fill = PatternFill("solid", fgColor=color)
            row += 1

        # === أعرض الأعمدة ===
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20

    def _create_issues_sheets(self, wb, seo_issues):
        """إنشاء ورقة لكل مستوى من المشاكل."""
        severity_emojis = {
            "🔴 Critical": ("🔴 Critical Issues", COLORS["critical"]),
            "🟠 High": ("🟠 High Priority", COLORS["high"]),
            "🟡 Medium": ("🟡 Medium Priority", COLORS["medium"]),
            "🟢 Low": ("🟢 Low Priority", COLORS["low"]),
        }

        for severity, (sheet_name, bg_color) in severity_emojis.items():
            issues = seo_issues.get("by_severity", {}).get(severity, [])
            if not issues:
                continue

            ws = wb.create_sheet(sheet_name)

            # Headers
            columns = ["category", "issue_type", "description", "affected_count", "recommendation"]
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
                cell.font = Font(bold=True, color=COLORS["header_text"])
                cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Data
            for row_idx, issue in enumerate(issues, start=2):
                for col_idx, col_name in enumerate(columns, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=issue.get(col_name, ""))
                    cell.fill = PatternFill("solid", fgColor=bg_color)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Column widths
            ws.column_dimensions["A"].width = 15  # category
            ws.column_dimensions["B"].width = 30  # issue_type
            ws.column_dimensions["C"].width = 50  # description
            ws.column_dimensions["D"].width = 15  # affected_count
            ws.column_dimensions["E"].width = 50  # recommendation

            # Freeze first row
            ws.freeze_panes = "A2"

    def _create_data_sheet(
        self,
        wb,
        sheet_name: str,
        data: list[dict],
        important_columns: list[str] = None,
        max_rows: int = 50000,
    ):
        """إنشاء ورقة بيانات عامة."""
        ws = wb.create_sheet(sheet_name)

        if not data:
            ws["A1"] = "لا توجد بيانات"
            return

        # تحديد الأعمدة
        if important_columns:
            # نأخذ المهمة أولاً، ثم الباقي
            all_keys = set()
            for item in data:
                all_keys.update(item.keys())
            columns = [c for c in important_columns if c in all_keys]
            other_columns = sorted(all_keys - set(columns))
            columns.extend(other_columns)
        else:
            # كل المفاتيح من أول صف
            columns = list(data[0].keys()) if data else []

        # Headers
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color=COLORS["header_text"])
            cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data (with limit)
        for row_idx, item in enumerate(data[:max_rows], start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = item.get(col_name, "")
                # تحويل lists/dicts إلى string
                if isinstance(value, (list, dict)):
                    value = str(value)[:32700]  # حد Excel للخلية
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit columns (مع حد أقصى)
        for col_idx, col_name in enumerate(columns, start=1):
            max_len = max(
                len(str(col_name)),
                max((len(str(item.get(col_name, ""))) for item in data[:1000]), default=10)
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 60)

        # Freeze header row + add auto-filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _to_dict(self, obj: Any) -> dict[str, Any]:
        """تحويل dataclass إلى dict."""
        if is_dataclass(obj):
            return asdict(obj)
        if isinstance(obj, dict):
            return obj
        return {"value": str(obj)}

    def _get_attr(self, obj: Any, attr: str, default: Any = None) -> Any:
        """جلب attribute من dataclass أو dict."""
        if isinstance(obj, dict):
            return obj.get(attr, default)
        return getattr(obj, attr, default)
